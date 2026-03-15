#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
filter_foods_preserve_format.py

Process EACH input Excel file independently, preserving its original columns.
For each file:
  - Detect name, carbs, protein, fat columns (with candidates list).
  - Group rows by exact food name.
  - Two-pass streaming:
      Pass 1: compute preliminary means per group for carbs/protein/fat.
      Pass 2: exclude rows deviating > threshold from those means (per macro, skip check if mean is 0/NaN or value is NaN);
              compute final means from remaining rows; choose a representative row (first passing row, otherwise first row) and
              replace its macro columns with the final means. Keep all other columns as in the representative row.
  - Save one output Excel per input with suffix "__mean{int(threshold*100)}".
  - Save a report CSV per input with counts per food name and global totals.

This script is memory-aware:
  - Uses openpyxl in read_only mode to iterate rows twice (streaming). It stores only small dictionaries per unique food name.

Usage:
  python filter_foods_preserve_format.py \
    --inputs "/mnt/data/20250327_가공식품DB_147999건.xlsx" \
             "/mnt/data/20250408_음식DB.xlsx" \
             "/mnt/data/국가표준식품성분표_250426공개.xlsx" \
    --threshold 0.2 \
    --name_cols 식품명 제품명 품목명 name \
    --carb_cols 탄수화물 "탄수화물(g)" carbs carbohydrate \
    --protein_cols 단백질 "단백질(g)" protein \
    --fat_cols 지방 "지방(g)" fat

Outputs (per input file, in the same directory as the input):
  <basename>__mean20.xlsx
  <basename>__mean20_report.csv
"""

import argparse
import os
import re
import sys
from collections import defaultdict, OrderedDict
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install openpyxl", file=sys.stderr)
    raise

try:
    import pandas as pd
except ImportError:
    print("ERROR: pandas is required. pip install pandas", file=sys.stderr)
    raise


def _norm_colname(x: Optional[str]) -> str:
    if x is None:
        return ""
    return re.sub(r"\s+", "", str(x)).lower()


def detect_columns(header_cells: List[str], candidates: List[str]) -> Optional[int]:
    norm_header = [_norm_colname(h) for h in header_cells]
    norm_cands = [_norm_colname(c) for c in candidates]
    for idx, h in enumerate(norm_header):
        for c in norm_cands:
            if c and (h == c or c in h or h in c):
                return idx
    return None


def parse_float(x) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x) if float(x) >= 0 else None
    s = str(x).strip()
    if not s or s.lower() in {"nan", "none", "null", "-"}:
        return None
    s = s.replace(",", "")
    # Drop unit-like chars
    s = re.sub(r"[a-zA-Zㄱ-ㅎ가-힣()%/㎖㎎㎍μgµgmgmlL]+", "", s)
    s = s.replace("±", "")
    try:
        v = float(s)
        return v if v >= 0 else None
    except ValueError:
        return None


def pass1_compute_means(ws, name_idx: int, carb_idx: Optional[int], protein_idx: Optional[int], fat_idx: Optional[int]) -> Dict[str, Dict[str, float]]:
    """Stream sheet and compute preliminary sums/counts per food name for macros."""
    sums = defaultdict(lambda: {"c": 0.0, "pc": 0, "p": 0.0, "pp": 0, "f": 0.0, "pf": 0})
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue

        if carb_idx is not None and carb_idx < len(row):
            c = parse_float(row[carb_idx])
            if c is not None:
                sums[name]["c"] += c
                sums[name]["pc"] += 1

        if protein_idx is not None and protein_idx < len(row):
            p = parse_float(row[protein_idx])
            if p is not None:
                sums[name]["p"] += p
                sums[name]["pp"] += 1

        if fat_idx is not None and fat_idx < len(row):
            f = parse_float(row[fat_idx])
            if f is not None:
                sums[name]["f"] += f
                sums[name]["pf"] += 1

    means = {}
    for name, acc in sums.items():
        means[name] = {
            "carb_mean": (acc["c"] / acc["pc"]) if acc["pc"] else None,
            "prot_mean": (acc["p"] / acc["pp"]) if acc["pp"] else None,
            "fat_mean":  (acc["f"] / acc["pf"]) if acc["pf"] else None,
        }
    return means


def within_threshold(val: Optional[float], mean_val: Optional[float], thr: float) -> bool:
    if val is None or mean_val is None or mean_val == 0:
        # if we cannot judge, don't exclude based on this column
        return True
    lower = mean_val * (1 - thr)
    upper = mean_val * (1 + thr)
    return (lower <= val <= upper)


def pass2_filter_and_aggregate(ws, header: List[str], name_idx: int, carb_idx: Optional[int], protein_idx: Optional[int], fat_idx: Optional[int],
                               means: Dict[str, Dict[str, float]], threshold: float
                               ) -> Tuple[List[OrderedDict], List[Dict]]:
    """
    Return:
      - rows_out: representative rows with original columns preserved; macros replaced by final means
      - report_rows: per-food counts
    """
    # Accumulators for final means after filtering
    acc_final = defaultdict(lambda: {"c": 0.0, "pc": 0, "p": 0.0, "pp": 0, "f": 0.0, "pf": 0,
                                     "total": 0, "used": 0, "rep_row": None, "rep_used": False})

    # First pass (again) to count total rows per name
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue
        acc_final[name]["total"] += 1

    # Second pass to filter and build rep rows
    # Need to re-open the sheet iterator (openpyxl generator is one-shot)
    ws.reset_dimensions()
    # Skip header row manually
    it = ws.iter_rows(min_row=2, values_only=True)

    for row in it:
        name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue

        # Get values
        c = parse_float(row[carb_idx]) if carb_idx is not None and carb_idx < len(row) else None
        p = parse_float(row[protein_idx]) if protein_idx is not None and protein_idx < len(row) else None
        f = parse_float(row[fat_idx]) if fat_idx is not None and fat_idx < len(row) else None

        m = means.get(name, {})
        ok_c = within_threshold(c, m.get("carb_mean"), threshold)
        ok_p = within_threshold(p, m.get("prot_mean"), threshold)
        ok_f = within_threshold(f, m.get("fat_mean"), threshold)

        use_row = ok_c and ok_p and ok_f

        if use_row:
            if c is not None:
                acc_final[name]["c"] += c
                acc_final[name]["pc"] += 1
            if p is not None:
                acc_final[name]["p"] += p
                acc_final[name]["pp"] += 1
            if f is not None:
                acc_final[name]["f"] += f
                acc_final[name]["pf"] += 1
            acc_final[name]["used"] += 1

            # Capture representative row (first used row wins)
            if acc_final[name]["rep_row"] is None:
                # preserve original columns order
                od = OrderedDict()
                for i, col in enumerate(header):
                    od[col] = row[i] if i < len(row) else None
                acc_final[name]["rep_row"] = od
                acc_final[name]["rep_used"] = True

        # If we don't have a representative yet, keep the first encountered row for backup
        if acc_final[name]["rep_row"] is None:
            od = OrderedDict()
            for i, col in enumerate(header):
                od[col] = row[i] if i < len(row) else None
            acc_final[name]["rep_row"] = od
            acc_final[name]["rep_used"] = False

    # Build outputs
    rows_out: List[OrderedDict] = []
    report_rows: List[Dict] = []

    for name, acc in acc_final.items():
        rep = acc["rep_row"]
        if rep is None:
            continue

        # Compute final means (fallback to preliminary mean if none used)
        final_c = (acc["c"] / acc["pc"]) if acc["pc"] else means.get(name, {}).get("carb_mean")
        final_p = (acc["p"] / acc["pp"]) if acc["pp"] else means.get(name, {}).get("prot_mean")
        final_f = (acc["f"] / acc["pf"]) if acc["pf"] else means.get(name, {}).get("fat_mean")

        # Replace macro columns if present
        if carb_idx is not None and carb_idx < len(header):
            rep[header[carb_idx]] = final_c
        if protein_idx is not None and protein_idx < len(header):
            rep[header[protein_idx]] = final_p
        if fat_idx is not None and fat_idx < len(header):
            rep[header[fat_idx]] = final_f

        rows_out.append(rep)

        report_rows.append({
            "식품명": name,
            "총개수": acc["total"],
            "사용개수": acc["used"],
            "제거개수": acc["total"] - acc["used"],
            "대표행이_필터통과여부": acc["rep_used"],
        })

    # 정렬 순서 수정용
    # rows_out.sort(key=lambda od: str(od.get(header[name_idx], "")) if name_idx is not None else "")
    report_rows.sort(key=lambda x: (-x["제거개수"], -x["총개수"], x["식품명"]))

    return rows_out, report_rows


def process_one_file(path: str, threshold: float,
                     name_cands: List[str], carb_cands: List[str], prot_cands: List[str], fat_cands: List[str],
                     sheet_name: Optional[str] = None) -> Tuple[str, str]:
    """Return (output_xlsx_path, report_csv_path)."""
    if not os.path.exists(path):
        print(f"[WARN] Missing file: {path}", file=sys.stderr)
        return "", ""

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # Header detection
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [h if h is not None else "" for h in header_row]
    name_idx = detect_columns(header, name_cands)
    carb_idx = detect_columns(header, carb_cands)
    prot_idx = detect_columns(header, prot_cands)
    fat_idx = detect_columns(header, fat_cands)

    if name_idx is None:
        wb.close()
        print(f"[WARN] {os.path.basename(path)}: cannot find a name column. Skipping.", file=sys.stderr)
        return "", ""

    # Pass 1: preliminary means
    means = pass1_compute_means(ws, name_idx, carb_idx, prot_idx, fat_idx)

    # Need to reopen workbook to reset streaming iterator
    wb.close()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # Skip header (we'll pass header separately)
    _ = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

    rows_out, report_rows = pass2_filter_and_aggregate(ws, header, name_idx, carb_idx, prot_idx, fat_idx, means, threshold)
    wb.close()

    # Write outputs next to input
    base_dir = os.path.dirname(path)
    base_name = os.path.splitext(os.path.basename(path))[0]
    tag = f"mean{int(threshold*100)}"
    out_xlsx = os.path.join(base_dir, f"{base_name}__{tag}.xlsx")
    rep_csv  = os.path.join(base_dir, f"{base_name}__{tag}_report.csv")

    # DataFrame preserving column order
    if rows_out:
        df = pd.DataFrame(rows_out, columns=header)
        try:
            df.to_excel(out_xlsx, index=False)
        except Exception as e:
            print(f"[WARN] Failed to write {out_xlsx}: {e}", file=sys.stderr)

    # Report
    if report_rows:
        rep_df = pd.DataFrame(report_rows)
        # Prepend global totals
        global_row = pd.DataFrame([{
            "식품명": "__GLOBAL__",
            "총개수": int(rep_df["총개수"].sum()),
            "사용개수": int(rep_df["사용개수"].sum()),
            "제거개수": int(rep_df["제거개수"].sum()),
            "대표행이_필터통과여부": None,
        }])
        rep_df = pd.concat([global_row, rep_df], ignore_index=True)
        rep_df.to_csv(rep_csv, index=False, encoding="utf-8-sig")

    print(f"[DONE] {os.path.basename(path)} -> {os.path.basename(out_xlsx)}, {os.path.basename(rep_csv)}")
    return out_xlsx, rep_csv


def main():
    ap = argparse.ArgumentParser(description="20% deviation filter + averaging per file, preserving original columns.")
    ap.add_argument("--inputs", nargs="+", required=True, help="Excel file paths")
    ap.add_argument("--threshold", type=float, default=0.2, help="Allowed relative deviation (default 0.2)")
    ap.add_argument("--sheet", default=None, help="Sheet name (default: first sheet)")

    ap.add_argument("--name_cols", nargs="+",
                    default=["식품명", "제품명", "품목명", "name", "foodname", "food", "품명", "메뉴명"])
    ap.add_argument("--carb_cols", nargs="+",
                    default=["탄수화물", "탄수화물(g)", "carbs", "carbohydrate", "탄수화물(가당)"])
    ap.add_argument("--protein_cols", nargs="+",
                    default=["단백질", "단백질(g)", "protein"])
    ap.add_argument("--fat_cols", nargs="+",
                    default=["지방", "지방(g)", "fat", "지방(총)"])

    args = ap.parse_args()

    for path in args.inputs:
        try:
            process_one_file(
                path=path,
                threshold=args.threshold,
                name_cands=args.name_cols,
                carb_cands=args.carb_cols,
                prot_cands=args.protein_cols,
                fat_cands=args.fat_cols,
                sheet_name=args.sheet
            )
        except Exception as e:
            print(f"[ERROR] Failed processing {path}: {e}", file=sys.stderr)


if __name__ == "__main__":
    main()
